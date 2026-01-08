"""
Template Loader Module

Reads and parses Template 1.xlsx to extract predefined scopes and categories.
Provides dynamic scope management instead of hardcoding.
"""

import logging
from typing import List, Dict, Optional
from pathlib import Path
import openpyxl

from src.config import settings

logger = logging.getLogger(__name__)


class TemplateLoader:
    """Loads scope definitions from Template 1.xlsx."""

    def __init__(self, template_path: Optional[str] = None):
        """
        Initialize template loader.

        Args:
            template_path: Path to Template 1.xlsx (optional, uses default if not provided)
        """
        if template_path is None:
            # Look for Template 1.xlsx in project root
            possible_paths = [
                Path(__file__).parent.parent / "Template 1.xlsx",
                Path(__file__).parent.parent / "template_1.xlsx",
                Path("Template 1.xlsx"),
            ]
            for path in possible_paths:
                if path.exists():
                    template_path = str(path)
                    break

        if not template_path or not Path(template_path).exists():
            raise FileNotFoundError(
                f"Template 1.xlsx not found. Searched: {possible_paths}"
            )

        self.template_path = Path(template_path)
        logger.info(f"Loading template from: {self.template_path}")

    def load_scopes(self) -> List[Dict[str, str]]:
        """
        Load scope definitions from Template 1.xlsx.

        Expected structure (flexible):
        - Sheet name: "Scopes" or "Policies" or first sheet
        - Columns: name, description, category (or similar)

        Returns:
            List of scope dictionaries with keys: name, description, category
        """
        try:
            workbook = openpyxl.load_workbook(self.template_path)

            # Try common sheet names first
            sheet_names = ["Scopes", "scopes", "Policies", "policies"]
            sheet = None
            for name in sheet_names:
                if name in workbook.sheetnames:
                    sheet = workbook[name]
                    break

            # Use first sheet if common names not found
            if sheet is None:
                sheet = workbook.active
                logger.info(f"Using sheet: {sheet.title}")
            else:
                logger.info(f"Using sheet: {sheet.title}")

            scopes = self._parse_scopes_sheet(sheet)

            workbook.close()
            logger.info(f"Loaded {len(scopes)} scopes from template")
            return scopes

        except Exception as e:
            logger.error(f"Error loading template: {e}")
            raise

    def _parse_scopes_sheet(self, sheet) -> List[Dict[str, str]]:
        """
        Parse scopes from worksheet.

        Flexible parsing to handle different column orders.

        Args:
            sheet: openpyxl worksheet

        Returns:
            List of scope dictionaries
        """
        scopes = []

        # Find header row (first row with data)
        headers = {}
        header_row_idx = None

        for idx, row in enumerate(sheet.iter_rows(max_row=10, values_only=False), 1):
            cells = [cell.value for cell in row]

            # Skip empty rows
            if all(v is None for v in cells):
                continue

            # Check if this looks like a header row
            header_candidates = [
                'name', 'scope', 'category', 'description',
                'type', 'label', 'field'
            ]

            cells_lower = [str(c).lower() if c else '' for c in cells]

            if any(h in ' '.join(cells_lower) for h in header_candidates):
                header_row_idx = idx
                # Map column positions to field names
                for col_idx, cell in enumerate(row):
                    if cell.value:
                        col_name = str(cell.value).lower().strip()
                        headers[col_idx] = col_name
                logger.info(f"Found headers at row {header_row_idx}: {headers}")
                break

        if not headers:
            logger.warning("Could not detect headers, assuming first data row is header")
            return self._parse_without_headers(sheet)

        # Parse data rows
        start_row = header_row_idx + 1

        for row_idx, row in enumerate(sheet.iter_rows(min_row=start_row, values_only=False), start_row):
            cells = [cell.value for cell in row]

            # Skip empty rows
            if all(v is None for v in cells):
                continue

            # Extract fields based on detected headers
            scope_data = self._extract_scope_data(cells, headers)

            if scope_data.get('name'):  # Only add if name exists
                scopes.append(scope_data)

        return scopes

    def _extract_scope_data(self, cells: List, headers: Dict[int, str]) -> Dict[str, str]:
        """
        Extract scope data from row based on detected headers.

        Args:
            cells: Row cells
            headers: Column index to header name mapping

        Returns:
            Dictionary with scope data
        """
        scope_data = {
            'name': None,
            'description': None,
            'category': None,
        }

        for col_idx, header in headers.items():
            if col_idx >= len(cells):
                continue

            value = cells[col_idx]
            if value is None:
                continue

            value = str(value).strip()

            # Map to scope fields
            if any(h in header for h in ['name', 'scope', 'label']):
                scope_data['name'] = value
            elif any(h in header for h in ['description', 'desc', 'detail']):
                scope_data['description'] = value
            elif any(h in header for h in ['category', 'type', 'group']):
                scope_data['category'] = value

        return scope_data

    def _parse_without_headers(self, sheet) -> List[Dict[str, str]]:
        """
        Parse scopes assuming standard column order (name, description, category).

        Args:
            sheet: openpyxl worksheet

        Returns:
            List of scope dictionaries
        """
        scopes = []

        for row_idx, row in enumerate(sheet.iter_rows(values_only=True, min_row=1), 1):
            if not row or all(v is None for v in row):
                continue

            # Skip header-like rows
            if row_idx == 1 and any(
                'name' in str(v).lower() or 'scope' in str(v).lower()
                for v in row if v
            ):
                continue

            name = str(row[0]).strip() if row and row[0] else None
            description = str(row[1]).strip() if len(row) > 1 and row[1] else None
            category = str(row[2]).strip() if len(row) > 2 and row[2] else None

            if name:
                scopes.append({
                    'name': name,
                    'description': description,
                    'category': category,
                })

        return scopes


def load_scopes_from_template(template_path: Optional[str] = None) -> List[Dict[str, str]]:
    """
    Convenience function to load scopes from Template 1.xlsx.

    Args:
        template_path: Optional path to template file

    Returns:
        List of scope dictionaries
    """
    loader = TemplateLoader(template_path)
    return loader.load_scopes()


def get_scope_names(template_path: Optional[str] = None) -> List[str]:
    """
    Get list of scope names from template.

    Args:
        template_path: Optional path to template file

    Returns:
        List of scope names
    """
    scopes = load_scopes_from_template(template_path)
    return [scope['name'] for scope in scopes]
